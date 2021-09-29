import openpyxl

#criar planilha (book)
book = openpyxl.Workbook()

#visualizar paginas existentes
print(book.sheetnames)

#criar pagina
book.create_sheet('Contatos')
del book['Sheet']

#selecionar pagina
contatos = book['Contatos']

head = ['Nome','Email','Telefone']
lines = [['João','joao@hotmail.com','55487896'],
    ['Maria','maria@hotmail.com','78456325'],
    ['José','jose@outlook.com','98654874'],
    ['Carlos','carlos@gmail.com','25556644'],
    ['Antonio','anton179@yahoo.com','00556685'],
    ['Fulano de Tal','alofulano@alo.com','99707070']]


#adicionando dados
contatos.append(head)

for line in lines:
    contatos.append(line)

#salvar planiha
book.save('Contatos.xlsx')

#fechar planilha
book.close()
print('#### planilha fechada #####')

#lendo a planilha aq
contacts = openpyxl.load_workbook('Contatos.xlsx')

listContatos = contacts['Contatos']

#imprimi os dados



for c in listContatos.iter_rows():
    for cell in c:
        print(cell.value)

print('#########################')

#outra forma
for c in listContatos.iter_rows():
    #o f é para fazer a formatação e separalos por ,
    print(f'{c[0].value}, {c[1].value}, {c[2].value}')


print('#########################')

#alterar o valor de uma celula
for c in listContatos.iter_rows():
    for cell in c:
        if cell.value == 'Fulano de Tal':
            cell.value = 'Não é Fulano é Beltrano'

#salva as alterações
contacts.save('Contatos2.xlsx')

for c in listContatos.iter_rows():
    #o f é para fazer a formatação e separalos por ,
    print(f'{c[0].value}, {c[1].value}, {c[2].value}')

